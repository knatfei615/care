"""Flask application – PICU pharmacy monitoring web interface."""

from __future__ import annotations

from io import BytesIO
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace

import openpyxl
from flask import Flask, after_this_request, jsonify, render_template, request, send_file
from flask_login import current_user, login_required
from flask_wtf.csrf import CSRFProtect, generate_csrf
from werkzeug.utils import secure_filename

from config import (
    DATA_DIR,
    FLASK_DEBUG,
    MAX_UPLOAD_MB,
    OPENAI_API_KEY,
    OPENAI_BASE_URL,
    OPENAI_MODEL,
    PORT,
    RECORD_TEMPLATE_PATH,
    SECRET_KEY,
    SQLALCHEMY_DATABASE_URI,
)
from excel_io import (
    add_patient,
    ExcelError,
    backup_workbook_identifiers,
    find_workbook,
    get_prior_note_context,
    get_slot_detail,
    get_slot_status_with_summary,
    list_patients,
    prepare_download_workbook,
    redact_workbook_identifiers,
    redact_workbooks,
    save_note,
)
from generate_picu_note import load_templates, render_note
from import_case_list import WorkbookPaths, run_import
from llm import structure_note
from models import db, init_db, login_manager

app = Flask(__name__)
csrf = CSRFProtect()
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_MB * 1024 * 1024
app.config["SECRET_KEY"] = SECRET_KEY
app.config["SQLALCHEMY_DATABASE_URI"] = SQLALCHEMY_DATABASE_URI

db.init_app(app)
login_manager.init_app(app)
csrf.init_app(app)

from admin import admin_bp  # noqa: E402
from auth import auth_bp  # noqa: E402

app.register_blueprint(auth_bp)
app.register_blueprint(admin_bp)


@app.context_processor
def inject_csrf_token():
    """Always expose csrf_token() in templates."""
    return {"csrf_token": generate_csrf}


def _user_data_dir() -> Path:
    """Return the data directory scoped to the current logged-in user."""
    return DATA_DIR / "users" / str(current_user.id)


def _ensure_user_dir() -> Path:
    d = _user_data_dir()
    d.mkdir(parents=True, exist_ok=True)
    redact_workbooks(d)
    return d


def _api_error(
    message: str,
    status: int = 400,
    error_type: str = "business_error",
    recovery_hint: str | None = None,
):
    payload: dict[str, str] = {"error": message, "error_type": error_type}
    if recovery_hint:
        payload["recovery_hint"] = recovery_hint
    return jsonify(payload), status


@app.route("/healthz")
def healthz():
    return jsonify(ok=True)


@app.route("/favicon.ico")
def favicon():
    return "", 204


# ── Pages ───────────────────────────────────────────────────────────

@app.route("/")
@login_required
def index():
    return render_template("index.html")


# ── Upload / Download ──────────────────────────────────────────────

@app.route("/api/upload", methods=["POST"])
@login_required
def upload():
    user_dir = _ensure_user_dir()
    file = request.files.get("file")
    if not file or not file.filename:
        return _api_error("未选择文件。", error_type="upload_error", recovery_hint="请选择一个 .xlsm 监护工作簿后重试。")
    filename = secure_filename(file.filename)
    if not filename or not filename.lower().endswith(".xlsm"):
        return _api_error("仅支持 .xlsm 文件。", error_type="upload_error", recovery_hint="请上传医院监护工作簿（.xlsm）。")

    dest = user_dir / filename
    file.save(str(dest))
    backup_workbook_identifiers(dest)
    redact_workbook_identifiers(dest)
    return jsonify(ok=True, filename=filename)


@app.route("/api/template/case-list")
@login_required
def download_case_list_template():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "病例清单模板"

    headers = [
        "住院号",
        "病人姓名",
        "年龄",
        "性别",
        "体重(kg)",
        "当前科室",
        "病区",
        "床号",
        "入院日期",
        "入院诊断",
    ]
    header_hints = {
        "住院号": "必填，如 123456",
        "病人姓名": "必填，如 张三",
        "年龄": "可填岁/月，如 2岁、6月",
        "性别": "男 或 女",
        "体重(kg)": "请填数字，如 12.5",
        "当前科室": "如：儿科",
        "病区": "如：PICU",
        "床号": "床位号，如 01",
        "入院日期": "建议 YYYY-MM-DD 格式",
        "入院诊断": "主要诊断",
    }

    sheet.append(headers)
    sheet.freeze_panes = "A2"

    for col_idx, title in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = openpyxl.styles.Font(bold=True)
        hint = header_hints.get(title)
        if hint:
            cell.comment = openpyxl.comments.Comment(hint, "系统模板")
        if title in {"入院诊断", "当前科室", "病区"}:
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 22
        else:
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 14

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    workbook.close()
    return send_file(
        buffer,
        as_attachment=True,
        download_name="病例清单模板.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/import-caselist", methods=["POST"])
@login_required
def import_caselist():
    user_dir = _ensure_user_dir()
    case_file = request.files.get("file")
    if not case_file or not case_file.filename:
        return _api_error("未选择病例清单文件。", error_type="upload_error", recovery_hint="请上传 .xlsx 的病例清单。")

    case_filename = secure_filename(case_file.filename)
    if not case_filename.lower().endswith(".xlsx"):
        return _api_error("仅支持 .xlsx 病例清单。", error_type="upload_error", recovery_hint="请确认文件扩展名为 .xlsx。")

    existing_wb = find_workbook(user_dir)
    template_wb = existing_wb
    if template_wb is None:
        if RECORD_TEMPLATE_PATH and RECORD_TEMPLATE_PATH.exists():
            template_wb = RECORD_TEMPLATE_PATH
        else:
            return _api_error(
                "未找到可用于导入的监护模板工作簿。",
                status=500,
                error_type="config_error",
                recovery_hint="请先上传一个 .xlsm 监护工作簿，或在环境变量 RECORD_TEMPLATE_PATH 配置默认模板路径。",
            )

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    tmp_case_path = user_dir / f"~caselist-{timestamp}.xlsx"
    output_wb = user_dir / f"住院药学诊查记录-{timestamp}-已导入.xlsm"
    case_file.save(str(tmp_case_path))

    args = SimpleNamespace(
        pharmacist=(request.form.get("pharmacist") or "").strip(),
        employee_id=(request.form.get("employee_id") or "").strip(),
        unit_mode=(request.form.get("unit_mode") or "both").strip(),
    )
    if args.unit_mode not in {"department", "ward", "both"}:
        args.unit_mode = "both"

    try:
        imported_count, _ = run_import(
            WorkbookPaths(
                case_list=tmp_case_path,
                record_book=template_wb,
                output_book=output_wb,
            ),
            args,
        )
        backup_workbook_identifiers(output_wb)
        redact_workbook_identifiers(output_wb)
    except SystemExit:
        return _api_error(
            "病例清单导入失败。",
            error_type="excel_error",
            recovery_hint="请检查病例清单列名是否完整（住院号、病人姓名、床号、入院诊断等）。",
        )
    except Exception as exc:
        return _api_error(
            f"导入失败：{exc}",
            error_type="excel_error",
            recovery_hint="请确认病例清单格式正确并重试。",
        )
    finally:
        tmp_case_path.unlink(missing_ok=True)

    return jsonify(
        ok=True,
        filename=output_wb.name,
        imported_count=imported_count,
        source_template=template_wb.name,
    )


@app.route("/api/download")
@login_required
def download():
    user_dir = _ensure_user_dir()
    wb = find_workbook(user_dir)
    if not wb:
        return _api_error("云端没有工作簿，请先上传。", status=404, error_type="upload_error", recovery_hint="先上传 .xlsm 或从病例清单创建工作簿。")
    try:
        export_wb = prepare_download_workbook(wb)
    except ExcelError as exc:
        return _api_error(str(exc), error_type="excel_error", recovery_hint="请重新上传原始工作簿后再下载。")

    @after_this_request
    def cleanup_export(response):
        try:
            export_wb.unlink(missing_ok=True)
        except OSError:
            pass
        return response

    return send_file(str(export_wb), as_attachment=True, download_name=wb.name)


# ── Patients ────────────────────────────────────────────────────────

@app.route("/api/patients")
@login_required
def patients():
    user_dir = _ensure_user_dir()
    wb = find_workbook(user_dir)
    if not wb:
        return _api_error("云端没有工作簿，请先上传。", status=404, error_type="upload_error", recovery_hint="上传 .xlsm 工作簿，或使用“从病例清单创建工作簿”。")
    try:
        rows = list_patients(wb)
    except ExcelError as exc:
        return _api_error(str(exc), error_type="excel_error", recovery_hint="请检查工作簿格式是否符合监护模板。")
    return jsonify(patients=rows)


@app.route("/api/patients", methods=["POST"])
@login_required
def add_patient_api():
    user_dir = _ensure_user_dir()
    wb = find_workbook(user_dir)
    if not wb:
        return _api_error("云端没有工作簿，请先上传。", status=404, error_type="upload_error", recovery_hint="请先上传工作簿再添加患者。")

    data = request.get_json(silent=True) or {}
    try:
        patient = add_patient(wb, data)
    except ExcelError as exc:
        return _api_error(str(exc), error_type="excel_error", recovery_hint="请检查住院号/姓名是否填写，且患者区是否未满。")

    return jsonify(ok=True, patient=patient)


# ── Generate (LLM) ─────────────────────────────────────────────────

@app.route("/api/generate", methods=["POST"])
@login_required
def generate():
    if not OPENAI_API_KEY:
        return _api_error("服务器未配置 OPENAI_API_KEY。", status=500, error_type="config_error", recovery_hint="请联系管理员配置 AI API Key。")

    data = request.get_json(silent=True) or {}
    raw_text = (data.get("raw_text") or "").strip()
    if not raw_text:
        return _api_error("请输入查房记录。", error_type="validation_error", recovery_hint="先填写查房口述，再点击 AI 生成。")

    patient_info = (data.get("patient_info") or "").strip()
    prior_notes = ""
    row_idx = data.get("row_idx")
    if row_idx not in (None, ""):
        user_dir = _ensure_user_dir()
        wb = find_workbook(user_dir)
        if not wb:
            return _api_error("云端没有工作簿，请先上传。", status=404, error_type="upload_error", recovery_hint="先上传 .xlsm 或从病例清单创建工作簿。")
        try:
            prior_notes = get_prior_note_context(wb, int(row_idx))
        except (TypeError, ValueError):
            return _api_error("患者行号格式错误。", error_type="validation_error", recovery_hint="请重新选择患者后再生成。")
        except ExcelError as exc:
            return _api_error(str(exc), error_type="excel_error", recovery_hint="请刷新页面后重新选择患者。")

    result = structure_note(
        OPENAI_API_KEY,
        OPENAI_MODEL,
        patient_info,
        raw_text,
        OPENAI_BASE_URL,
        prior_notes=prior_notes,
    )
    return jsonify(result)


@app.route("/api/templates")
@login_required
def templates():
    try:
        specs = load_templates()
    except SystemExit:
        return _api_error("模板文件加载失败。", status=500, error_type="config_error", recovery_hint="请检查 picu_note_templates.json 是否存在且格式正确。")

    payload = [
        {
            "id": t.template_id,
            "name": t.name,
            "description": t.description,
            "default_note_type": t.default_note_type,
            "fields": [{"key": f.key, "label": f.label, "default": f.default} for f in t.fields],
        }
        for t in specs
    ]
    return jsonify(templates=payload)


@app.route("/api/templates/render", methods=["POST"])
@login_required
def template_render():
    data = request.get_json(silent=True) or {}
    template_id = (data.get("template_id") or "").strip()
    values = data.get("values") or {}
    if not template_id:
        return _api_error("缺少 template_id。", error_type="validation_error", recovery_hint="请先选择模板。")
    if not isinstance(values, dict):
        return _api_error("模板参数格式错误。", error_type="validation_error", recovery_hint="请刷新页面后重试。")

    try:
        specs = load_templates()
    except SystemExit:
        return _api_error("模板文件加载失败。", status=500, error_type="config_error", recovery_hint="请检查模板配置文件。")

    target = next((t for t in specs if t.template_id == template_id), None)
    if not target:
        return _api_error("模板不存在。", status=404, error_type="validation_error", recovery_hint="请重新选择模板。")

    field_values = {
        f.key: str(values.get(f.key, "")).strip() or f.default
        for f in target.fields
    }
    try:
        note = render_note(target, field_values, multiline=False)
    except KeyError as exc:
        return _api_error(
            f"模板配置错误：缺少占位符 {exc}",
            status=500,
            error_type="config_error",
            recovery_hint="请联系管理员检查模板配置。",
        )
    except (ValueError, IndexError) as exc:
        return _api_error(
            f"模板渲染失败：{exc}",
            error_type="template_error",
            recovery_hint="请检查填入内容是否包含特殊字符（如大括号 {}），修改后重试。",
        )
    return jsonify(
        ok=True,
        note=note,
        default_note_type=target.default_note_type,
    )


# ── Save to Excel ──────────────────────────────────────────────────

@app.route("/api/save", methods=["POST"])
@login_required
def save():
    user_dir = _ensure_user_dir()
    wb = find_workbook(user_dir)
    if not wb:
        return jsonify(error="云端没有工作簿，请先上传。"), 404

    data = request.get_json(silent=True) or {}
    row_idx = data.get("row_idx")
    note_text = (data.get("note_text") or "").strip()
    level = data.get("level", "")
    note_type = data.get("note_type", "")
    date_str = data.get("date", "")
    overwrite_slot = data.get("overwrite_slot")
    force_overwrite = bool(data.get("force_overwrite"))

    if not row_idx or not note_text:
        return _api_error("缺少必填项（row_idx, note_text）。", error_type="validation_error", recovery_hint="请先选择患者并填写监护意见。")
    if not level:
        return _api_error("缺少监护级别。", error_type="validation_error", recovery_hint="请先选择一级/二级/三级监护。")
    if not note_type:
        return _api_error("缺少监护类型。", error_type="validation_error", recovery_hint="请先选择药学查房/药学监护等类型。")
    if not date_str:
        return _api_error("缺少日期。", error_type="validation_error", recovery_hint="请先选择记录日期。")
    if len(note_text) < 20:
        return _api_error("监护意见过短。", error_type="validation_error", recovery_hint="建议补充完整的四段式内容后再保存。")

    skip_structure_check = bool(data.get("skip_structure_check"))
    if not skip_structure_check:
        required_markers = ("主观资料：", "客观资料：", "分析评估：", "药学监护建议：")
        if not all(marker in note_text for marker in required_markers):
            return _api_error("监护意见缺少四段结构。", error_type="validation_error", recovery_hint="建议包含“主观资料/客观资料/分析评估/药学监护建议”四段后再保存。")

    try:
        record_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return _api_error("日期格式错误，应为 YYYY-MM-DD。", error_type="validation_error", recovery_hint="请从日期选择器重新选择日期。")

    if overwrite_slot:
        try:
            detail = get_slot_detail(wb, int(row_idx), int(overwrite_slot))
        except ExcelError as exc:
            return _api_error(str(exc), error_type="excel_error", recovery_hint="请刷新后重试，或重新选择患者。")
        if detail.get("has_note") and not force_overwrite:
            existing_note = (detail.get("note") or "").strip()
            return jsonify(
                ok=False,
                needs_confirm=True,
                overwrite_slot=int(overwrite_slot),
                old_preview=existing_note[:80],
                new_preview=note_text[:80],
                message=f"记录{overwrite_slot}已有内容，确认覆盖吗？",
            )

    try:
        result = save_note(wb, int(row_idx), record_date, level, note_type, note_text, overwrite_slot)
    except ExcelError as exc:
        return _api_error(str(exc), error_type="excel_error", recovery_hint="请检查记录槽位状态，必要时选择覆盖槽位。")

    return jsonify(ok=True, **result)


# ── Slot status ─────────────────────────────────────────────────────

@app.route("/api/slots/<int:row_idx>")
@login_required
def slots(row_idx):
    user_dir = _ensure_user_dir()
    wb = find_workbook(user_dir)
    if not wb:
        return _api_error("云端没有工作簿，请先上传。", status=404, error_type="upload_error")
    try:
        payload = get_slot_status_with_summary(wb, row_idx)
    except ExcelError as exc:
        return _api_error(str(exc), error_type="excel_error")
    return jsonify(slots=payload["slots"], summary=payload["summary"])


@app.route("/api/slots/<int:row_idx>/<int:slot_index>")
@login_required
def slot_detail(row_idx, slot_index):
    user_dir = _ensure_user_dir()
    wb = find_workbook(user_dir)
    if not wb:
        return _api_error("云端没有工作簿，请先上传。", status=404, error_type="upload_error")
    try:
        detail = get_slot_detail(wb, row_idx, slot_index)
    except ExcelError as exc:
        return _api_error(str(exc), error_type="excel_error")
    return jsonify(slot=detail)


# ── User info API (for frontend) ───────────────────────────────────

@app.route("/api/me")
@login_required
def me():
    return jsonify(
        username=current_user.username,
        display_name=current_user.display_name,
        role=current_user.role,
    )


# ── Main ────────────────────────────────────────────────────────────

DATA_DIR.mkdir(parents=True, exist_ok=True)
init_db(app)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=PORT, debug=FLASK_DEBUG)
