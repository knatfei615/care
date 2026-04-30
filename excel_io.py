"""Excel I/O wrapper for the web layer.

Re-uses functions from the existing CLI scripts while converting
``SystemExit`` (raised by ``fail()``) into a catchable ``ExcelError``.
"""

from __future__ import annotations

from io import BytesIO
import json
import threading
from contextlib import suppress
from datetime import date, datetime
from pathlib import Path
import re
from typing import Any

import openpyxl

from import_case_list import RECORD_KEYWORD, TEMP_PREFIX
from write_picu_note_to_excel import (
    NOTE_SLOTS,
    LEVEL_OPTIONS,
    MAX_ROW,
    START_ROW,
    TYPE_OPTIONS,
    PatientRow,
    format_cell,
    load_patient_rows,
    find_previous_defaults,
    resolve_slot,
    write_note_to_sheet,
)

_lock = threading.Lock()
ID_COLUMN = "D"
NAME_COLUMN = "F"
BASE_COLUMNS = ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K")
ANON_ID_PREFIX = "ANON-"
TEMP_EXPORT_PREFIX = "~export-"
ANON_NAME_PREFIX = "匿名患者"
STATUS_TODAY_UPDATED = "today_updated"
STATUS_PENDING = "pending"
STATUS_STALE_48H = "stale_48h"
STATUS_NO_RECORD = "no_record"

DIAGNOSIS_TAG_RULES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("抗菌药物", ("感染", "抗感染", "抗菌", "肺炎", "败血症", "sepsis", "abx", "antibiotic")),
    ("TDM", ("tdm", "血药浓度", "万古霉素", "他克莫司", "丙戊酸", "地高辛")),
    ("CRRT", ("crrt", "透析", "血液净化", "连续肾脏替代", "ecmo")),
    ("肾功能异常", ("肾功能", "肾损伤", "少尿", "无尿", "肌酐", "bun", "aki", "ckd")),
    ("ADR", ("不良反应", "过敏", "皮疹", "肝损伤", "adr", "药物不良", "药疹")),
)

class ExcelError(Exception):
    """Raised when an Excel operation fails."""


def _catch(fn, *args, **kwargs):
    """Call *fn* and convert ``SystemExit`` into ``ExcelError``."""
    try:
        return fn(*args, **kwargs)
    except SystemExit as exc:
        raise ExcelError(str(exc)) from exc


def _close_workbook(book) -> None:
    """Close an openpyxl workbook and archives opened by keep_vba."""
    for attr in ("_archive", "vba_archive"):
        archive = getattr(book, attr, None)
        if archive:
            with suppress(Exception):
                archive.close()
    book.close()


def _identifier_backup_path(wb_path: Path) -> Path:
    return wb_path.with_name(f"{wb_path.stem}.identifiers.json")


def _medications_path(wb_path: Path) -> Path:
    return wb_path.with_name(f"{wb_path.stem}.medications.json")


def _load_medications_data(wb_path: Path) -> dict[str, Any]:
    path = _medications_path(wb_path)
    if not path.exists():
        return {"rows": {}}

    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {"rows": {}}

    rows = data.get("rows")
    if not isinstance(rows, dict):
        rows = {}
    return {"rows": rows}


def get_patient_medications(wb_path: Path, row_idx: int) -> dict[str, str]:
    """Return medication-order sidecar data for one patient row."""
    row_key = str(row_idx)
    with _lock:
        data = _load_medications_data(wb_path)
        row = data.get("rows", {}).get(row_key, {})

    if not isinstance(row, dict):
        row = {}
    return {
        "medications": format_cell(row.get("medications")),
        "updated_at": format_cell(row.get("updated_at")),
    }


def set_patient_medications(wb_path: Path, row_idx: int, medications: str) -> dict[str, str]:
    """Persist current medication orders for one patient row in a sidecar file."""
    medication_text = format_cell(medications).strip()[:2000]
    updated_at = datetime.now().isoformat(timespec="seconds")
    row_key = str(row_idx)

    with _lock:
        data = _load_medications_data(wb_path)
        rows = data.setdefault("rows", {})
        if not isinstance(rows, dict):
            rows = {}
            data["rows"] = rows

        rows[row_key] = {
            "medications": medication_text,
            "updated_at": updated_at,
        }

        _medications_path(wb_path).write_text(
            json.dumps({"rows": rows}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    return {"medications": medication_text, "updated_at": updated_at}


def load_all_medications(wb_path: Path) -> dict[int, str]:
    """Return all saved medication-order text by patient row index."""
    with _lock:
        data = _load_medications_data(wb_path)
        rows = data.get("rows", {})

    result: dict[int, str] = {}
    if not isinstance(rows, dict):
        return result

    for row_idx_text, values in rows.items():
        if not isinstance(values, dict):
            continue
        medications = format_cell(values.get("medications"))
        if not medications:
            continue
        try:
            result[int(row_idx_text)] = medications
        except (TypeError, ValueError):
            continue
    return result


def _parse_excel_date(value: Any) -> datetime | None:
    """Parse excel/date-like values into datetime."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    text = format_cell(value).strip()
    if not text:
        return None

    text = text.replace("年", "-").replace("月", "-").replace("日", "")
    text = text.replace("/", "-").strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def _extract_tags_from_diagnosis(diagnosis: str) -> list[str]:
    normalized = (diagnosis or "").lower()
    tags: list[str] = []
    for tag, keywords in DIAGNOSIS_TAG_RULES:
        if any(keyword.lower() in normalized for keyword in keywords):
            tags.append(tag)
    return tags


def _extract_major_issue(note_text: str) -> str:
    if not note_text:
        return ""

    match = re.search(r"主观资料：\s*(.*?)\s*(?:客观资料：|$)", note_text, flags=re.S)
    if not match:
        match = re.search(r"问题：\s*(.*?)\s*(?:分析：|$)", note_text, flags=re.S)
    if not match:
        return ""

    issue = re.sub(r"\s+", " ", match.group(1)).strip("；;。 ")
    return issue[:100]


def _collect_slot_records(sheet, row_idx: int) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Collect raw slot status list and effective note records."""
    slots: list[dict[str, Any]] = []
    records: list[dict[str, Any]] = []

    for slot in NOTE_SLOTS:
        note_val = format_cell(sheet[f"{slot['note']}{row_idx}"].value)
        date_raw = sheet[f"{slot['date']}{row_idx}"].value
        date_val = format_cell(date_raw)
        level_val = format_cell(sheet[f"{slot['level']}{row_idx}"].value)
        type_val = format_cell(sheet[f"{slot['type']}{row_idx}"].value)
        parsed_date = _parse_excel_date(date_raw or date_val)

        slots.append({
            "index": slot["index"],
            "has_note": bool(note_val),
            "date": date_val,
            "preview": note_val[:60] if note_val else "",
            "level": level_val,
            "note_type": type_val,
        })

        if note_val:
            records.append({
                "slot_index": slot["index"],
                "date": date_val,
                "parsed_date": parsed_date,
                "level": level_val,
                "note_type": type_val,
                "note": note_val,
            })

    records.sort(
        key=lambda item: (
            item["parsed_date"] or datetime.min,
            item["slot_index"],
        ),
        reverse=True,
    )
    return slots, records


def _compute_status(last_note_dt: datetime | None) -> str:
    if last_note_dt is None:
        return STATUS_NO_RECORD

    now = datetime.now()
    if last_note_dt.date() == now.date():
        return STATUS_TODAY_UPDATED

    elapsed = now - last_note_dt
    if elapsed.total_seconds() > 48 * 3600:
        return STATUS_STALE_48H
    return STATUS_PENDING


def _build_tracking_summary(records: list[dict[str, Any]]) -> dict[str, Any]:
    latest = records[0] if records else None
    latest_dt = latest["parsed_date"] if latest else None
    latest_date = ""
    if latest:
        latest_date = latest["date"] or (latest_dt.strftime("%Y-%m-%d") if latest_dt else "")

    recent_records = []
    for item in records[:3]:
        preview = re.sub(r"\s+", " ", item["note"]).strip()[:50]
        recent_records.append({
            "slot_index": item["slot_index"],
            "date": item["date"] or (item["parsed_date"].strftime("%Y-%m-%d") if item["parsed_date"] else ""),
            "level": item["level"],
            "note_type": item["note_type"],
            "preview": preview,
        })

    return {
        "latest_note_date": latest_date,
        "major_issue": _extract_major_issue(latest["note"]) if latest else "",
        "recent_records": recent_records,
    }


def _format_prior_note_record(index: int, record: dict[str, Any]) -> str:
    """Format one prior note record for LLM context."""
    date_text = record["date"]
    if not date_text and record["parsed_date"]:
        date_text = record["parsed_date"].strftime("%Y-%m-%d")

    note_text = re.sub(r"\s+", " ", record["note"]).strip()
    return (
        f"记录{index}："
        f"日期：{date_text or '未填写'}；"
        f"分级：{record['level'] or '未填写'}；"
        f"类型：{record['note_type'] or '未填写'}；"
        f"内容：{note_text}"
    )


def get_prior_note_context(wb_path: Path, row_idx: int, limit: int = 6) -> str:
    """Return formatted prior note context for LLM generation."""
    if limit <= 0:
        return ""

    with _lock:
        book = openpyxl.load_workbook(BytesIO(wb_path.read_bytes()), keep_vba=True, data_only=True)
        try:
            sheet = book[book.sheetnames[0]]
            _load_patient_row(sheet, row_idx)
            _, records = _collect_slot_records(sheet, row_idx)
        finally:
            _close_workbook(book)

    return "\n".join(
        _format_prior_note_record(index, record)
        for index, record in enumerate(records[:limit], start=1)
    )


def _migrated_identifier_backup_path(backup_path: Path) -> Path:
    migrated_path = backup_path.with_name(f"{backup_path.name}.migrated")
    if not migrated_path.exists():
        return migrated_path

    for index in range(1, 1000):
        candidate = backup_path.with_name(f"{backup_path.name}.migrated.{index}")
        if not candidate.exists():
            return candidate
    raise ExcelError("无法生成 identifiers 迁移备份文件名。")


def restore_anonymized_workbook(wb_path: Path) -> bool:
    """Restore old row-based aliases from the identifier sidecar if present."""
    backup_path = _identifier_backup_path(wb_path)
    if not backup_path.exists():
        return False

    try:
        backup_data = json.loads(backup_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return False

    rows = backup_data.get("rows", {})
    if not isinstance(rows, dict):
        return False

    changed = False
    with _lock:
        book = openpyxl.load_workbook(BytesIO(wb_path.read_bytes()), keep_vba=True)
        try:
            sheet = book[book.sheetnames[0]]
            for row_idx_text, values in rows.items():
                if not isinstance(values, dict):
                    continue
                try:
                    row_idx = int(row_idx_text)
                except (TypeError, ValueError):
                    continue
                if row_idx < START_ROW or row_idx > MAX_ROW:
                    continue

                alias_num = row_idx - START_ROW + 1
                expected_id = f"{ANON_ID_PREFIX}{alias_num:04d}"
                expected_name = f"{ANON_NAME_PREFIX}{alias_num:03d}"
                current_id = format_cell(sheet[f"{ID_COLUMN}{row_idx}"].value)
                current_name = format_cell(sheet[f"{NAME_COLUMN}{row_idx}"].value)
                real_id = format_cell(values.get("inpatient_no"))
                real_name = format_cell(values.get("name"))

                if current_id == expected_id and real_id:
                    sheet[f"{ID_COLUMN}{row_idx}"] = real_id
                    changed = True
                if current_name == expected_name and real_name:
                    sheet[f"{NAME_COLUMN}{row_idx}"] = real_name
                    changed = True

            if changed:
                book.save(wb_path)
        finally:
            _close_workbook(book)

    backup_path.rename(_migrated_identifier_backup_path(backup_path))
    return changed


# ── workbook discovery ──────────────────────────────────────────────

def find_workbook(data_dir: Path) -> Path | None:
    """Return the first ``.xlsm`` workbook found in *data_dir*, or ``None``."""
    if not data_dir.is_dir():
        return None
    candidates = [
        p for p in data_dir.iterdir()
        if p.is_file()
        and p.suffix.lower() == ".xlsm"
        and not p.name.startswith(TEMP_PREFIX)
        and not p.name.startswith(TEMP_EXPORT_PREFIX)
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


# ── patient list ────────────────────────────────────────────────────

def list_patients(wb_path: Path) -> list[dict]:
    """Return a JSON-friendly list of patients with previous defaults."""
    with _lock:
        book = openpyxl.load_workbook(BytesIO(wb_path.read_bytes()), keep_vba=True, data_only=True)
        try:
            sheet = book[book.sheetnames[0]]
            patients = load_patient_rows(sheet)
            result = []
            for p in patients:
                prev_level, prev_type = find_previous_defaults(sheet, p)
                _, records = _collect_slot_records(sheet, p.row_idx)
                latest = records[0] if records else None
                latest_dt = latest["parsed_date"] if latest else None
                last_note_date = ""
                if latest:
                    last_note_date = latest["date"] or (latest_dt.strftime("%Y-%m-%d") if latest_dt else "")
                status = _compute_status(latest_dt)
                result.append({
                    "row_idx": p.row_idx,
                    "inpatient_no": p.inpatient_no,
                    "bed_no": p.bed_no,
                    "name": p.name,
                    "age": p.age,
                    "sex": p.sex,
                    "weight": p.weight,
                    "admission_date": p.admission_date,
                    "diagnosis": p.diagnosis,
                    "prev_level": prev_level,
                    "prev_type": prev_type,
                    "last_note_date": last_note_date,
                    "status": status,
                    "tags": _extract_tags_from_diagnosis(p.diagnosis),
                })
        finally:
            _close_workbook(book)
    return result


def build_ai_patient_info(wb_path: Path, row_idx: int) -> str:
    """Return patient context for the LLM without name or inpatient number."""
    with _lock:
        book = openpyxl.load_workbook(BytesIO(wb_path.read_bytes()), keep_vba=True, data_only=True)
        try:
            sheet = book[book.sheetnames[0]]
            patient = _load_patient_row(sheet, row_idx)
        finally:
            _close_workbook(book)

    parts = [
        f"床号：{patient.bed_no}" if patient.bed_no else "",
        f"年龄：{patient.age}" if patient.age else "",
        f"性别：{patient.sex}" if patient.sex else "",
        f"体重：{patient.weight}kg" if patient.weight else "",
        f"入院日期：{patient.admission_date}" if patient.admission_date else "",
        f"入院诊断：{patient.diagnosis}" if patient.diagnosis else "",
    ]
    return "，".join(part for part in parts if part)


def _clean_patient_payload(payload: dict[str, Any]) -> dict[str, str]:
    """Normalize manual patient payload from request JSON."""
    return {
        "A": format_cell(payload.get("department")),
        "B": format_cell(payload.get("pharmacist")),
        "C": format_cell(payload.get("employee_id")),
        "D": format_cell(payload.get("inpatient_no")),
        "E": format_cell(payload.get("bed_no")),
        "F": format_cell(payload.get("name")),
        "G": format_cell(payload.get("age")),
        "H": format_cell(payload.get("sex")),
        "I": format_cell(payload.get("weight")),
        "J": format_cell(payload.get("admission_date")),
        "K": format_cell(payload.get("diagnosis")),
    }


def add_patient(wb_path: Path, payload: dict[str, Any]) -> dict[str, Any]:
    """Append one patient base-info row into A-K columns."""
    row_data = _clean_patient_payload(payload)
    if not row_data["D"] and not row_data["F"]:
        raise ExcelError("住院号和姓名不能同时为空。")

    target_row: int | None = None
    with _lock:
        book = openpyxl.load_workbook(BytesIO(wb_path.read_bytes()), keep_vba=True)
        try:
            sheet = book[book.sheetnames[0]]

            for row_idx in range(START_ROW, MAX_ROW + 1):
                inpatient_no = format_cell(sheet[f"D{row_idx}"].value)
                name = format_cell(sheet[f"F{row_idx}"].value)
                if not inpatient_no and not name:
                    target_row = row_idx
                    break

            if target_row is None:
                raise ExcelError(f"患者区域已满（{START_ROW}-{MAX_ROW} 行）。")

            for column in BASE_COLUMNS:
                sheet[f"{column}{target_row}"] = row_data[column]

            book.save(wb_path)
        finally:
            _close_workbook(book)

    return {
        "row_idx": target_row,
        "department": row_data["A"],
        "pharmacist": row_data["B"],
        "employee_id": row_data["C"],
        "inpatient_no": row_data["D"],
        "bed_no": row_data["E"],
        "name": row_data["F"],
        "age": row_data["G"],
        "sex": row_data["H"],
        "weight": row_data["I"],
        "admission_date": row_data["J"],
        "diagnosis": row_data["K"],
    }


# ── slot status ─────────────────────────────────────────────────────

def _load_patient_row(sheet, row_idx: int) -> PatientRow:
    """Load and validate patient row data from worksheet."""
    if row_idx < START_ROW or row_idx > MAX_ROW:
        raise ExcelError(f"行号超出范围（{START_ROW}-{MAX_ROW}）。")

    patient = PatientRow(
        row_idx=row_idx,
        inpatient_no=format_cell(sheet[f"D{row_idx}"].value),
        bed_no=format_cell(sheet[f"E{row_idx}"].value),
        name=format_cell(sheet[f"F{row_idx}"].value),
        age=format_cell(sheet[f"G{row_idx}"].value),
        sex=format_cell(sheet[f"H{row_idx}"].value),
        weight=format_cell(sheet[f"I{row_idx}"].value),
        admission_date=format_cell(sheet[f"J{row_idx}"].value),
        diagnosis=format_cell(sheet[f"K{row_idx}"].value),
    )

    if not patient.inpatient_no and not patient.name:
        raise ExcelError(f"行 {row_idx} 没有患者数据。")

    return patient


def get_slot_status(wb_path: Path, row_idx: int) -> list[dict]:
    """Return the occupancy status of all 6 note slots for a given row."""
    with _lock:
        book = openpyxl.load_workbook(BytesIO(wb_path.read_bytes()), keep_vba=True, data_only=True)
        try:
            sheet = book[book.sheetnames[0]]
            _load_patient_row(sheet, row_idx)
            result, _ = _collect_slot_records(sheet, row_idx)
        finally:
            _close_workbook(book)
    return result


def get_slot_status_with_summary(wb_path: Path, row_idx: int) -> dict[str, Any]:
    """Return slot status with patient tracking summary for a row."""
    with _lock:
        book = openpyxl.load_workbook(BytesIO(wb_path.read_bytes()), keep_vba=True, data_only=True)
        try:
            sheet = book[book.sheetnames[0]]
            _load_patient_row(sheet, row_idx)
            slots, records = _collect_slot_records(sheet, row_idx)
            summary = _build_tracking_summary(records)
            summary["status"] = _compute_status(records[0]["parsed_date"] if records else None)
        finally:
            _close_workbook(book)
    return {"slots": slots, "summary": summary}


def get_slot_detail(wb_path: Path, row_idx: int, slot_index: int) -> dict:
    """Return the full content of one note slot for a given patient row."""
    slot = next((s for s in NOTE_SLOTS if s["index"] == slot_index), None)
    if not slot:
        raise ExcelError("记录槽位必须为 1-6。")

    with _lock:
        book = openpyxl.load_workbook(BytesIO(wb_path.read_bytes()), keep_vba=True, data_only=True)
        try:
            sheet = book[book.sheetnames[0]]
            _load_patient_row(sheet, row_idx)

            note_val = format_cell(sheet[f"{slot['note']}{row_idx}"].value)
            date_val = format_cell(sheet[f"{slot['date']}{row_idx}"].value)
            level_val = format_cell(sheet[f"{slot['level']}{row_idx}"].value)
            type_val = format_cell(sheet[f"{slot['type']}{row_idx}"].value)
        finally:
            _close_workbook(book)

    return {
        "index": slot["index"],
        "has_note": bool(note_val),
        "date": date_val,
        "level": level_val,
        "note_type": type_val,
        "note": note_val,
    }


# ── save note ───────────────────────────────────────────────────────

def save_note(
    wb_path: Path,
    row_idx: int,
    record_date: date,
    level: str,
    note_type: str,
    note_text: str,
    overwrite_slot: int | None = None,
) -> dict:
    """Write a note into the workbook and return a summary dict."""
    with _lock:
        book = openpyxl.load_workbook(BytesIO(wb_path.read_bytes()), keep_vba=True)
        try:
            sheet = book[book.sheetnames[0]]
            patient = _load_patient_row(sheet, row_idx)

            slot = _catch(resolve_slot, sheet, patient, overwrite_slot)

            write_note_to_sheet(
                sheet=sheet,
                patient=patient,
                slot=slot,
                record_date=record_date,
                level=level,
                note_type=note_type,
                note=note_text,
            )
            book.save(wb_path)
        finally:
            _close_workbook(book)
    return {
        "row_idx": row_idx,
        "name": patient.name,
        "slot_index": slot["index"],
        "date": record_date.isoformat(),
        "level": level,
        "note_type": note_type,
    }
