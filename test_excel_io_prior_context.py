import tempfile
import unittest
import json
from io import BytesIO
from pathlib import Path

import openpyxl

from excel_io import (
    build_ai_patient_info,
    get_prior_note_context,
    get_patient_medications,
    load_all_medications,
    restore_anonymized_workbook,
    set_patient_medications,
)


class PriorNoteContextTest(unittest.TestCase):
    def test_formats_existing_notes_in_reverse_chronological_order(self):
        with tempfile.TemporaryDirectory() as tmp:
            wb_path = Path(tmp) / "records.xlsm"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet["D3"] = "12345"
            sheet["F3"] = "测试患儿"

            sheet["L3"] = "2026-04-27"
            sheet["M3"] = "二级监护"
            sheet["N3"] = "药学查房"
            sheet["O3"] = "主观资料：昨日仍有发热。客观资料：CRP升高。分析评估：感染控制欠佳。药学监护建议：关注抗感染疗效。"

            sheet["P3"] = "2026-04-28"
            sheet["Q3"] = "一级监护"
            sheet["R3"] = "药学监护"
            sheet["S3"] = "主观资料：今日体温下降。客观资料：炎症指标下降。分析评估：治疗有效。药学监护建议：继续监测肾功能。"

            workbook.save(wb_path)
            workbook.close()

            context = get_prior_note_context(wb_path, row_idx=3)

        self.assertIn("记录1：日期：2026-04-28；分级：一级监护；类型：药学监护；内容：", context)
        self.assertIn("记录2：日期：2026-04-27；分级：二级监护；类型：药学查房；内容：", context)
        self.assertLess(context.index("2026-04-28"), context.index("2026-04-27"))
        self.assertNotIn("记录3：", context)

    def test_returns_empty_string_when_patient_has_no_prior_notes(self):
        with tempfile.TemporaryDirectory() as tmp:
            wb_path = Path(tmp) / "records.xlsm"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet["D3"] = "12345"
            sheet["F3"] = "测试患儿"
            workbook.save(wb_path)
            workbook.close()

            context = get_prior_note_context(wb_path, row_idx=3)

        self.assertEqual(context, "")


class PatientMedicationsTest(unittest.TestCase):
    def test_saves_trims_and_loads_medication_sidecar(self):
        with tempfile.TemporaryDirectory() as tmp:
            wb_path = Path(tmp) / "records.xlsm"

            saved = set_patient_medications(wb_path, 3, "  万古霉素 1g q12h IV  ")
            one_row = get_patient_medications(wb_path, 3)
            all_rows = load_all_medications(wb_path)

        self.assertEqual(saved["medications"], "万古霉素 1g q12h IV")
        self.assertTrue(saved["updated_at"])
        self.assertEqual(one_row["medications"], "万古霉素 1g q12h IV")
        self.assertEqual(all_rows, {3: "万古霉素 1g q12h IV"})

    def test_missing_medication_sidecar_returns_empty_payload(self):
        with tempfile.TemporaryDirectory() as tmp:
            wb_path = Path(tmp) / "records.xlsm"

            one_row = get_patient_medications(wb_path, 3)
            all_rows = load_all_medications(wb_path)

        self.assertEqual(one_row, {"medications": "", "updated_at": ""})
        self.assertEqual(all_rows, {})


class PatientIdentifierMigrationTest(unittest.TestCase):
    def test_build_ai_patient_info_excludes_name_and_inpatient_number(self):
        with tempfile.TemporaryDirectory() as tmp:
            wb_path = Path(tmp) / "records.xlsm"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet["D3"] = "12345"
            sheet["E3"] = "PICU-01"
            sheet["F3"] = "测试患儿"
            sheet["G3"] = "3岁"
            sheet["H3"] = "男"
            sheet["I3"] = "14"
            sheet["J3"] = "2026-04-30"
            sheet["K3"] = "重症肺炎"
            workbook.save(wb_path)
            workbook.close()

            patient_info = build_ai_patient_info(wb_path, row_idx=3)

        self.assertIn("床号：PICU-01", patient_info)
        self.assertIn("年龄：3岁", patient_info)
        self.assertIn("入院诊断：重症肺炎", patient_info)
        self.assertNotIn("12345", patient_info)
        self.assertNotIn("测试患儿", patient_info)
        self.assertNotIn("住院号", patient_info)
        self.assertNotIn("姓名", patient_info)

    def test_restore_anonymized_workbook_restores_exact_old_aliases(self):
        with tempfile.TemporaryDirectory() as tmp:
            wb_path = Path(tmp) / "records.xlsm"
            backup_path = Path(tmp) / "records.identifiers.json"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet["D3"] = "ANON-0001"
            sheet["F3"] = "匿名患者001"
            workbook.save(wb_path)
            workbook.close()
            backup_path.write_text(
                json.dumps({"rows": {"3": {"inpatient_no": "12345", "name": "测试患儿"}}}, ensure_ascii=False),
                encoding="utf-8",
            )

            changed = restore_anonymized_workbook(wb_path)
            restored = openpyxl.load_workbook(BytesIO(wb_path.read_bytes()))
            sheet = restored.active

            self.assertTrue(changed)
            self.assertEqual(sheet["D3"].value, "12345")
            self.assertEqual(sheet["F3"].value, "测试患儿")
            self.assertFalse(backup_path.exists())
            self.assertTrue(Path(tmp, "records.identifiers.json.migrated").exists())
            restored.close()

    def test_restore_anonymized_workbook_does_not_overwrite_non_alias_values(self):
        with tempfile.TemporaryDirectory() as tmp:
            wb_path = Path(tmp) / "records.xlsm"
            backup_path = Path(tmp) / "records.identifiers.json"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet["D3"] = "USER-EDITED-ID"
            sheet["F3"] = "用户已改名"
            workbook.save(wb_path)
            workbook.close()
            backup_path.write_text(
                json.dumps({"rows": {"3": {"inpatient_no": "12345", "name": "测试患儿"}}}, ensure_ascii=False),
                encoding="utf-8",
            )

            changed = restore_anonymized_workbook(wb_path)
            restored = openpyxl.load_workbook(BytesIO(wb_path.read_bytes()))
            sheet = restored.active

            self.assertFalse(changed)
            self.assertEqual(sheet["D3"].value, "USER-EDITED-ID")
            self.assertEqual(sheet["F3"].value, "用户已改名")
            self.assertFalse(backup_path.exists())
            self.assertTrue(Path(tmp, "records.identifiers.json.migrated").exists())
            restored.close()


if __name__ == "__main__":
    unittest.main()
