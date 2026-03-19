from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


CASE_KEYWORD = "病例清单"
RECORD_KEYWORD = "住院药学诊查记录"
TEMP_PREFIX = "~$"
START_ROW = 3
MAX_IMPORT_ROWS = 960


@dataclass
class WorkbookPaths:
    case_list: Path
    record_book: Path
    output_book: Path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Import base patient information from the latest case list workbook "
            "into the latest inpatient pharmacy monitoring workbook."
        )
    )
    parser.add_argument(
        "--case-list",
        type=Path,
        help="Path to the case list workbook. Defaults to the latest *病例清单*.xlsx file.",
    )
    parser.add_argument(
        "--record-book",
        type=Path,
        help="Path to the monitoring workbook. Defaults to the latest *住院药学诊查记录*.xlsm file.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Output workbook path. Defaults to '<record name>-已导入.xlsm'.",
    )
    parser.add_argument(
        "--pharmacist",
        default="",
        help="Default value for column B (临床药师).",
    )
    parser.add_argument(
        "--employee-id",
        default="",
        help="Default value for column C (工号).",
    )
    parser.add_argument(
        "--unit-mode",
        choices=("department", "ward", "both"),
        default="both",
        help="How to fill column A (科室/病区). Default: both.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Allow overwriting the output file if it already exists.",
    )
    return parser.parse_args()


def fail(message: str) -> None:
    print(f"[ERROR] {message}", file=sys.stderr)
    raise SystemExit(1)


def newest_matching_file(directory: Path, suffix: str, keyword: str) -> Path:
    candidates = [
        path
        for path in directory.iterdir()
        if path.is_file()
        and path.suffix.lower() == suffix
        and not path.name.startswith(TEMP_PREFIX)
        and keyword in path.stem
    ]
    if not candidates:
        fail(f"Could not find a file matching '*{keyword}*{suffix}' in {directory}")
    preferred = [path for path in candidates if "已导入" not in path.stem]
    target_pool = preferred or candidates
    return max(target_pool, key=lambda path: path.stat().st_mtime)


def resolve_paths(args: argparse.Namespace) -> WorkbookPaths:
    workdir = Path.cwd()
    case_list = args.case_list or newest_matching_file(workdir, ".xlsx", CASE_KEYWORD)
    record_book = args.record_book or newest_matching_file(workdir, ".xlsm", RECORD_KEYWORD)

    if not case_list.exists():
        fail(f"Case list workbook not found: {case_list}")
    if not record_book.exists():
        fail(f"Monitoring workbook not found: {record_book}")

    output_book = args.output
    if output_book is None:
        output_book = record_book.with_name(f"{record_book.stem}-已导入{record_book.suffix}")

    if output_book.exists() and not args.overwrite:
        fail(
            f"Output workbook already exists: {output_book}. "
            "Use --overwrite or provide a new --output path."
        )

    return WorkbookPaths(
        case_list=case_list.resolve(),
        record_book=record_book.resolve(),
        output_book=output_book.resolve(),
    )


def header_map(sheet: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in sheet[1]:
        if cell.value is None:
            continue
        headers[str(cell.value).strip()] = cell.column
    return headers


def normalize_value(value: Any) -> Any:
    if value is None:
        return ""
    return value


def build_unit_value(department: Any, ward: Any, mode: str) -> str:
    department_text = str(department).strip() if department not in (None, "") else ""
    ward_text = str(ward).strip() if ward not in (None, "") else ""

    if mode == "department":
        return department_text
    if mode == "ward":
        return ward_text
    if department_text and ward_text and department_text != ward_text:
        return f"{department_text} / {ward_text}"
    return department_text or ward_text


def extract_rows(
    case_sheet: openpyxl.worksheet.worksheet.Worksheet,
    pharmacist: str,
    employee_id: str,
    unit_mode: str,
) -> list[dict[str, Any]]:
    headers = header_map(case_sheet)
    required_headers = [
        "住院号",
        "病人姓名",
        "年龄",
        "性别",
        "体重(kg)",
        "当前科室",
        "病区",
        "床号",
        "入院日期",
        "入院诊断",
    ]
    missing = [name for name in required_headers if name not in headers]
    if missing:
        fail(f"Case list is missing required columns: {', '.join(missing)}")

    imported: list[dict[str, Any]] = []
    seen_keys: set[str] = set()
    for row_idx in range(2, case_sheet.max_row + 1):
        inpatient_no = normalize_value(case_sheet.cell(row_idx, headers["住院号"]).value)
        patient_name = normalize_value(case_sheet.cell(row_idx, headers["病人姓名"]).value)
        if inpatient_no == "" and patient_name == "":
            continue

        dedupe_key = f"{inpatient_no}|{patient_name}"
        if dedupe_key in seen_keys:
            continue
        seen_keys.add(dedupe_key)

        department = normalize_value(case_sheet.cell(row_idx, headers["当前科室"]).value)
        ward = normalize_value(case_sheet.cell(row_idx, headers["病区"]).value)

        imported.append(
            {
                "A": build_unit_value(department, ward, unit_mode),
                "B": pharmacist,
                "C": employee_id,
                "D": inpatient_no,
                "E": normalize_value(case_sheet.cell(row_idx, headers["床号"]).value),
                "F": patient_name,
                "G": normalize_value(case_sheet.cell(row_idx, headers["年龄"]).value),
                "H": normalize_value(case_sheet.cell(row_idx, headers["性别"]).value),
                "I": normalize_value(case_sheet.cell(row_idx, headers["体重(kg)"]).value),
                "J": normalize_value(case_sheet.cell(row_idx, headers["入院日期"]).value),
                "K": normalize_value(case_sheet.cell(row_idx, headers["入院诊断"]).value),
            }
        )

    if not imported:
        fail("No patient rows were found in the case list workbook.")
    if len(imported) > MAX_IMPORT_ROWS:
        fail(
            f"Case list has {len(imported)} rows, which exceeds the template capacity "
            f"of {MAX_IMPORT_ROWS} data rows."
        )
    return imported


def clear_base_columns(record_sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for row_idx in range(START_ROW, START_ROW + MAX_IMPORT_ROWS):
        for column in "ABCDEFGHIJK":
            record_sheet[f"{column}{row_idx}"] = None


def clear_note_columns(record_sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Clear all 6 note slots (columns L–AI, i.e. cols 12–35) so that stale
    monitoring records from a reused workbook don't leak into the new import."""
    for row_idx in range(START_ROW, START_ROW + MAX_IMPORT_ROWS):
        for col_idx in range(12, 36):
            record_sheet.cell(row=row_idx, column=col_idx).value = None


def write_rows(
    record_sheet: openpyxl.worksheet.worksheet.Worksheet,
    imported_rows: list[dict[str, Any]],
) -> None:
    for offset, row_data in enumerate(imported_rows):
        row_idx = START_ROW + offset
        for column, value in row_data.items():
            record_sheet[f"{column}{row_idx}"] = value


def run_import(paths: WorkbookPaths, args: argparse.Namespace) -> tuple[int, str]:
    case_book = openpyxl.load_workbook(paths.case_list, data_only=True)
    record_book = openpyxl.load_workbook(paths.record_book, keep_vba=True)

    case_sheet = case_book[case_book.sheetnames[0]]
    record_sheet = record_book[record_book.sheetnames[0]]

    imported_rows = extract_rows(
        case_sheet=case_sheet,
        pharmacist=args.pharmacist,
        employee_id=args.employee_id,
        unit_mode=args.unit_mode,
    )
    clear_base_columns(record_sheet)
    clear_note_columns(record_sheet)
    write_rows(record_sheet, imported_rows)

    record_book.save(paths.output_book)
    return len(imported_rows), record_sheet.title


def main() -> None:
    args = parse_args()
    paths = resolve_paths(args)
    imported_count, sheet_name = run_import(paths, args)

    print(f"[OK] Imported {imported_count} patient rows.")
    print(f"     Case list : {paths.case_list}")
    print(f"     Template  : {paths.record_book}")
    print(f"     Output    : {paths.output_book}")
    print(f"     Sheet     : {sheet_name}")
    if args.pharmacist:
        print(f"     Pharmacist: {args.pharmacist}")
    if args.employee_id:
        print(f"     EmployeeID: {args.employee_id}")
    print(f"     Unit mode : {args.unit_mode}")


if __name__ == "__main__":
    main()
